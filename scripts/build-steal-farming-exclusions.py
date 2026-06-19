import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEET_NAME = "除外リスト"
DEFAULT_SOURCE_FILENAME = "盗み金策_実測検証テンプレ_除外リスト追加.xlsx"
HEADER_ALIASES = {
    "monster_name": ["モンスター名", "monster_name"],
    "exclude_reason": ["除外理由", "exclude_reason"],
    "category": ["分類", "category"],
    "exclude_level": ["除外レベル", "exclude_level"],
    "judged_at": ["判定日", "judged_at"],
    "reviewer": ["確認者", "reviewer"],
    "site_exclude": ["サイト除外に反映", "site_exclude"],
    "memo": ["メモ", "memo"],
}
OUTPUT_HEADERS = list(HEADER_ALIASES.keys())


def normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(worksheet):
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [normalize_cell(value) for value in row]
        if "モンスター名" in values and "サイト除外に反映" in values:
            return row_number, values
    raise ValueError("除外リストのヘッダー行が見つかりません。")


def find_column_index(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def build_rows(source_path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{SHEET_NAME} シートが見つかりません。")

    worksheet = workbook[SHEET_NAME]
    header_row_number, headers = find_header_row(worksheet)
    column_indexes = {
        key: find_column_index(headers, aliases)
        for key, aliases in HEADER_ALIASES.items()
    }

    missing_required = [
        key
        for key in ("monster_name", "site_exclude")
        if column_indexes[key] is None
    ]
    if missing_required:
        raise ValueError(f"必須列が見つかりません: {', '.join(missing_required)}")

    rows = []
    for row in worksheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        monster_name_index = column_indexes["monster_name"]
        monster_name = normalize_cell(row[monster_name_index] if monster_name_index < len(row) else "")
        if not monster_name:
            continue

        output_row = {}
        for key, index in column_indexes.items():
            output_row[key] = normalize_cell(row[index] if index is not None and index < len(row) else "")
        rows.append(output_row)

    return rows


def main():
    repo_root = Path(__file__).resolve().parents[1]
    if len(sys.argv) >= 2:
        source_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        source_path = repo_root / "data" / DEFAULT_SOURCE_FILENAME
        if not source_path.exists():
            raise SystemExit(
                "Usage: python scripts/build-steal-farming-exclusions.py <xlsx-path> [output-csv]"
            )

    output_path = (
        Path(sys.argv[2]).expanduser().resolve()
        if len(sys.argv) >= 3
        else repo_root / "data" / "steal_farming_exclusions.csv"
    )

    rows = build_rows(source_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"wrote {len(rows)} rows: {output_path}")


if __name__ == "__main__":
    main()
